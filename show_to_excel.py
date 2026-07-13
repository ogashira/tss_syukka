import openpyxl
import openpyxl.styles # style用に追加import
from openpyxl.utils import get_column_letter
from typing import cast
from IExcel_output import IExcelOutput, AllPackings



class ShowToExcel:
    def __init__(self, book_name: str, allPackings: IExcelOutput) -> None:
        self._book_name = book_name
        # IExcelOutput型からAllPackings型にキャストする
        self._allPackings = cast(AllPackings, allPackings)

    def show_to_excel(self) -> None:
        wb = openpyxl.load_workbook(self._book_name)
        ws = wb["Sheet1"] # シートの取得
        lastRow: int = ws.max_row
        lastCol: int = ws.max_column

        ws.cell(row=1, column= lastCol + 1).value = '売価' # type: ignore
        line = openpyxl.styles.Side(style="medium", color="000000") # 太線・黒色
        myborder = openpyxl.styles.Border(bottom=line) # lineを上下左右すべてに適用

        # 売価を入力
        self._allPackings.show_uriKin_for_excel(ws, myborder)
        # 新しい列のアルファベット（例: 'E'）を取得
        target_col_letter = get_column_letter(lastCol + 1)
        # 「売価」の列幅だけを12に設定
        ws.column_dimensions[target_col_letter].width = 12

        wb.save(self._book_name)





