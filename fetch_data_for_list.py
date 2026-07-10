import sys
import csv
import platform
from  datetime import date, timedelta
from re import I
import warnings
import pandas as pd
from typing import List, Any, Tuple
from abc import ABC, abstractmethod
from get_idx import GetIdx
import openpyxl

warnings.filterwarnings('ignore', category=UserWarning)

def create_minYM_maxYM(syukka_date: str)-> Tuple:
    '''
    minM: 出荷日の月の前月
    maxM: 出荷日の月の翌月
    '''
    syukkaY = syukka_date[:4]
    syukkaM = syukka_date[4:6]
    if syukkaM == "12":
        minM = str(int(syukkaM) - 1)
        minY = syukkaY
        maxM = "01"
        maxY = str(int(syukkaY) + 1)
    elif syukkaM == "01":
        minM = "12"
        minY = str(int(syukkaY) - 1)
        maxM = str(int(syukkaM) + 1)
        maxY = syukkaY
    else:
        minM = str(int(syukkaM) - 1)
        minY = syukkaY
        maxM = str(int(syukkaM) + 1)
        maxY = syukkaY

    if len(maxM) == 1:
        maxM = "0" + maxM
    if len(minM) == 1:
        minM = "0" + minM

    maxYM = "'" + maxY + maxM + "'"
    minYM = "'" + minY + minM + "'"

    return minYM, maxYM



class IFetchDataForList(ABC):

    @abstractmethod
    def fetch_data(self)-> Tuple[List[str],List[List[Any]]]:
        pass


class FetchUriageSumiForPacking(IFetchDataForList):

    def __init__(self, cnxn, syukka_date:str) -> None:
        self.cnxn = cnxn
        self._syukka_date = "'" + syukka_date + "'"
        

    def fetch_data(self) -> Tuple[List[str],List[List[Any]]]:

        cursor = self.cnxn.cursor()

        sqlQuery = ("SELECT RURIDT.RurTokCD AS '得意先コード',"
                    " RURIDT.RurNonyuCD AS '納入先コード',"
                    " RURIHD.RurNonyuNam1 AS '納入先名称１',"
                    " RURIDT.RurCMNo AS '得意先注文ＮＯ',"
                    " RURIHD.RurUnsCD AS 'unsouCD',"
                    " MA_UNS.aitNam1 AS '依頼先',"
                    " RURIDT.RurUriDay AS '出荷日',"
                    " RURIDT.RurNODay AS '納期',"
                    " RURIDT.RurHinCD AS '売り品番',"
                    " RURIDT.RurHinNam AS '品名',"
                    " RURIDT.RurKoSu AS 'uriKosu',"
                    " RURIDT.RurKanriTniCD AS '受注単位',"
                    " RURMEI_U2002.RmeMHinCD AS 'motoHinCD',"
                    " RURMEI_U2002.RmeMSu AS 'motoSu',"
                    " RURMEI_U2002.RmeMtniCD AS 'motoTni',"
                    " RURIDT.RurUriTnk AS 'uriTnk',"
                    " RURIDT.RurUriKin AS 'uriKin',"
                    " RURIDT.RurMBiko AS '備考',"
                    " RURIDT.RurKojFrom AS 'factory_name',"
                    " RJYUCD.RjcFree1 AS 'add'"
                    " From dbo.RURIDT"
                    " JOIN dbo.RURIHD"
                    " ON RURIDT.RurUNo = RURIHD.RurUNo" 
                    " LEFT JOIN dbo.RURMEI_U2002"
                    " ON RURIDT.RurUNo = RURMEI_U2002.RmeUNo"
                    " AND RURIDT.RurUGNo = RURMEI_U2002.RmeUGNo"
                    " LEFT JOIN(" 
                        " SELECT MAITEM.AitCD1, MAITEM.AitNam1" 
                        " FROM dbo.MAITEM"
                        " WHERE MAITEM.AitAitKBN = 'A'" # A = 運送屋
                    ")MA_UNS ON RURIHD.RurUnsCD = MA_UNS.AitCD1"
                    " LEFT JOIN dbo.RJYUCD"
                    " ON RURIDT.RurJCNo = RJYUCD.RjcJCNo" 
                    " AND RURIDT.RurJGNo = RJYUCD.RjcJGNo"
                    " WHERE RURIDT.RurUriDay =" + self._syukka_date +
                    " AND RURIDT.RurTokCD < 'T6000'"
                    " AND RURIDT.RurTokCD <> 'T0000'"
                    " ORDER BY RURIDT.RurTokCD, RURIDT.RurNonyuCD, RURIDT.RurCMNo"
                    )

        data_list: List[List[Any]] = []
        cursor.execute(sqlQuery)

        # 1. カラム名を取得（リスト内包表記）
        # cursor.description は (名前, 型, 表示サイズ, ...) というタプルのリスト
        columns = [column[0] for column in cursor.description]

        # 4. 2次元リストへ変換
        # fetchall() はタプルのリストを返すため、リスト内包表記で各行をリスト化します
        try:
            data_list = [list(row) for row in cursor.fetchall()]
        except Exception as e:
            raise Exception(f'データベースfetch中に予期せぬエラーです FetchUriageSumiForPacking') from e
        finally:
            cursor.close()
            # cnxnは呼び出しもとでクローズ


        return columns, data_list


class FetchUriageSumi(IFetchDataForList):

    def __init__(self, cnxn, syukka_date:str) -> None:
        self.cnxn = cnxn
        self._syukka_date = "'" + syukka_date + "'"
        

    def fetch_data(self) -> Tuple[List[str],List[List[Any]]]:

        cursor = self.cnxn.cursor()

        sqlQuery = ("SELECT RURIDT.RurTokCD AS '得意先コード',"
                    " RURIDT.RurNonyuCD AS '納入先コード',"
                    " RURIHD.RurUnsCD AS 'unsou_code',"
                    " MA_UNS.AitNam1 AS 'unsou',"
                    " RURIDT.RurFreeKBN1 AS 'kubun_no',"
                    " KBN.KbnNam AS 'kubun',"
                    " RURIDT.RurUriDay AS '出荷予定日',"
                    " RURIDT.RurHinCD AS 'hinban',"
                    " RURIDT.RurHinNam AS '品名',"
                    " RURMEI.RmeLotNo AS 'lot',"
                    " RURMEI.RmeKoSu AS '受注数量'," 
                    " RURIDT.RurKanriTniCD AS '受注単位',"
                    " RURIHD.RurNonyuNam1 AS '納入先名称１',"
                    " RURIDT.RurCMNo AS '得意先注文ＮＯ',"
                    " RURIDT.RurMBiko AS '備考',"
                    " RJYUCD.RjcFree1 AS 'add',"  
                    " RURMEI.RmeKojFrom AS 'factory_name',"
                    " MA_NONYU.AitRyaku AS '納入先名',"
                    " RURMEI_U2002.RmeMHinCD AS 'motoHinCD',"
                    " RURMEI_U2002.RmeMtniCD AS 'motoTni',"
                    " RURMEI_U2002.RmeMSu AS  '振替元数量'" 
                    " From dbo.RURIDT"
                    " JOIN dbo.RURMEI"
                    " ON RURIDT.RurUNo = RURMEI.RmeUNo" 
                    " AND RURIDT.RurUGNo = RURMEI.RmeUGNo"
                    " JOIN dbo.RURIHD"
                    " ON RURIDT.RurUNo = RURIHD.RurUNo"
                    " LEFT JOIN dbo.RURMEI_U2002"
                    " ON RURIDT.RurUNo = RURMEI_U2002.RmeUNo"
                    " AND RURIDT.RurUGNo = RURMEI_U2002.RmeUGNo"
                    " AND RURMEI.RmeSeqNo = RURMEI_U2002.RmeSeqNo"
                    " LEFT JOIN dbo.MAITEM AS MA_NONYU"
                    " ON RURIDT.RurTokCD = MA_NONYU.AitCD1"
                    " AND RURIDT.RurNonyuCD = MA_NONYU.AitCD2"
                    " LEFT JOIN(" 
                        " SELECT MAITEM.AitCD1, MAITEM.AitNam1" 
                        " FROM dbo.MAITEM"
                        " WHERE MAITEM.AitAitKBN = 'A'" # A = 運送屋
                    ")MA_UNS ON RURIHD.RurUnsCD = MA_UNS.AitCD1"
                    " LEFT JOIN("
                        " SELECT MKUBUN.KbnCD, MKUBUN.KbnNam"  
                        " FROM dbo.MKUBUN"
                        " WHERE MKUBUN.KbnKBN = 'V'" # V = 配送区分
                    ")KBN ON RURIDT.RurFreeKBN1 = KBN.KbnCD"
                    " LEFT JOIN dbo.RJYUCD"
                    " ON RURIDT.RurJCNo = RJYUCD.RjcJCNo" 
                    " AND RURIDT.RurJGNo = RJYUCD.RjcJGNo"
                    " WHERE RURIDT.RurUriDay =" + self._syukka_date +
                    " AND RURIDT.RurTokCD < 'T6000'"
                    " AND RURIDT.RurTokCD <> 'T0000'"
                    " ORDER BY RURIDT.RurTokCD, RURIDT.RurNonyuCD"
                    )

        data_list: List[List[Any]] = []
        cursor.execute(sqlQuery)

        # 1. カラム名を取得（リスト内包表記）
        # cursor.description は (名前, 型, 表示サイズ, ...) というタプルのリスト
        columns = [column[0] for column in cursor.description]

        # 4. 2次元リストへ変換
        # fetchall() はタプルのリストを返すため、リスト内包表記で各行をリスト化します
        try:
            data_list = [list(row) for row in cursor.fetchall()]
        except Exception as e:
            raise Exception (f'データベースfetch中に予期せぬエラーです FetchUriageSumi') from e
        finally:
            cursor.close()
            # cnxnは呼び出しもとでクローズ

        return columns, data_list


class FetchTyuzan(IFetchDataForList):

    def __init__(self, cnxn, syukka_date:str) -> None:
        self.cnxn = cnxn
        self._syukka_date = syukka_date


    def fetch_data(self) -> Tuple[List[str],List[List[Any]]]:

        cursor = self.cnxn.cursor()

        sqlQuery = ("SELECT RjcJCNo AS '受注No',"
                    " RjcJGNo AS '受注行No',"
                    " RjcTokCD AS '得意先CD',"
                    " RjcNonyuCD AS '納入先CD',"
                    " RjcHinCD AS '受注品番',"
                    " RjcJcSu AS '受注数',"
                    " RjcURSu AS '売り数'"
                    " FROM RJYUCD"
                    " WHERE RjcSKDay =" + self._syukka_date +
                    " AND RjcTokCD < 'T6000'"
                    " AND RjcJcSu - RjcURSu > 0 "
                    " ORDER BY RjcJCNo, RjcJGNo"
                    )

        data_list: List[List[Any]] = []
        cursor.execute(sqlQuery)

        # 1. カラム名を取得（リスト内包表記）
        # cursor.description は (名前, 型, 表示サイズ, ...) というタプルのリスト
        columns = [column[0] for column in cursor.description]

        # 4. 2次元リストへ変換
        # fetchall() はタプルのリストを返すため、リスト内包表記で各行をリスト化します
        try:
            data_list = [list(row) for row in cursor.fetchall()]
        except Exception as e:
            raise Exception(f'データベースfetch中に予期せぬエラーです FetchTyuzan') from e
        finally:
            cursor.close()
            # cnxnは呼び出しもとでクローズ


        return columns, data_list


class FetchCalenderUnsouya(IFetchDataForList):

    def __init__(self, cnxn, syukka_date:str) -> None:
        self.cnxn = cnxn
        self._minYM, self._maxYM = create_minYM_maxYM(syukka_date)

    def fetch_data(self) -> Tuple[List[str],List[List[Any]]]:

        cursor = self.cnxn.cursor()

        sqlQuery = ("SELECT CalYM AS 'YYYYMM',"
                    " CalDay AS 'DD',"
                    " CalYobiJ AS 'Week',"
                    " CalFlg AS 'holiday'"
                    " FROM MCALEN"
                    " WHERE CalKojCD = '@0001'" 
                    " AND CalBuCD = 'DUMMY'"
                    " AND CalYM >=" + self._minYM +
                    " AND CalYM <=" + self._maxYM +
                    " ORDER BY CalYM, CalDay"
                    )

        data_list: List[List[Any]] = []
        cursor.execute(sqlQuery)

        # 1. カラム名を取得（リスト内包表記）
        # cursor.description は (名前, 型, 表示サイズ, ...) というタプルのリスト
        columns = [column[0] for column in cursor.description]

        # 4. 2次元リストへ変換
        # fetchall() はタプルのリストを返すため、リスト内包表記で各行をリスト化します
        try:
            data_list = [list(row) for row in cursor.fetchall()]
        except Exception as e:
            raise Exception(f'データベースfetch中に予期せぬエラーです FetchCalenderUnsouya') from e
        finally:
            cursor.close()
            # cnxnは呼び出しもとでクローズ


        return columns, data_list


class FetchCalenderToyo(IFetchDataForList):

    def __init__(self, cnxn, syukka_date:str) -> None:
        self.cnxn = cnxn

        self._minYM, self._maxYM = create_minYM_maxYM(syukka_date)


    def fetch_data(self) -> Tuple[List[str],List[List[Any]]]:

        cursor = self.cnxn.cursor()

        sqlQuery = ("SELECT CalYM AS 'YYYYMM',"
                    " CalDay AS 'DD',"
                    " CalYobiJ AS 'Week',"
                    " CalFlg AS 'holiday'"
                    " FROM MCALEN"
                    " WHERE CalKojCD = '@@@@@'" 
                    " AND CalYM >=" + self._minYM +
                    " AND CalYM <=" + self._maxYM +
                    " ORDER BY CalYM, CalDay"
                    )

        data_list: List[List[Any]] = []
        cursor.execute(sqlQuery)

        # 1. カラム名を取得（リスト内包表記）
        # cursor.description は (名前, 型, 表示サイズ, ...) というタプルのリスト
        columns = [column[0] for column in cursor.description]

        # 4. 2次元リストへ変換
        # fetchall() はタプルのリストを返すため、リスト内包表記で各行をリスト化します
        try:
            data_list = [list(row) for row in cursor.fetchall()]
        except Exception as e:
            raise Exception(f'データベースfetch中に予期せぬエラーです FetchCalenderToyo') from e
        finally:
            cursor.close()
            # cnxnは呼び出しもとでクローズ

        return columns, data_list


class FetchProductCan(IFetchDataForList):
    # PSマスタの品目１がGK(缶)のやつ

    def __init__(self, cnxn) -> None:
        self.cnxn = cnxn
        

    def fetch_data(self) -> Tuple[List[str],List[List[Any]]]:

        cursor = self.cnxn.cursor()

        sqlQuery = ("SELECT MPSMST.PsmHinCDO AS '親品番',"
                    " MPSMST.PsmHinCDK AS '子品番'"
                    " From dbo.MPSMST"
                    " LEFT JOIN dbo.MHINCD"
                    " ON MPSMST.PsmHinCDK = MHINCD.HinHinCD" 
                    " WHERE MHINCD.HinMokCD1 = 'GK'"
                    " ORDER BY MPSMST.PsmHinCDO"
                    )

        data_list: List[List[Any]] = []
        cursor.execute(sqlQuery)

        # 1. カラム名を取得（リスト内包表記）
        # cursor.description は (名前, 型, 表示サイズ, ...) というタプルのリスト
        columns = [column[0] for column in cursor.description]

        # 4. 2次元リストへ変換
        # fetchall() はタプルのリストを返すため、リスト内包表記で各行をリスト化します
        try:
            data_list = [list(row) for row in cursor.fetchall()]
        except Exception as e:
            raise Exception(f'データベースfetch中に予期せぬエラーです FetchProductCan') from e
        finally:
            cursor.close()
            # cnxnは呼び出しもとでクローズ

        return columns, data_list


class FetchTnju(IFetchDataForList):

    def __init__(self, cnxn) -> None:
        self.cnxn = cnxn
        

    def fetch_data(self) -> Tuple[List[str],List[List[Any]]]:

        cursor = self.cnxn.cursor()

        sqlQuery = ("SELECT MHINCD.HinHinCD AS 'hinban',"
                    " MHINCD.HinTju AS 'tnju'"
                    " From dbo.MHINCD"
                    " WHERE MHINCD.HinTniCD = 'CN'"
                    " ORDER BY MHINCD.HinHinCD"
                    )

        data_list: List[List[Any]] = []
        cursor.execute(sqlQuery)

        # 1. カラム名を取得（リスト内包表記）
        # cursor.description は (名前, 型, 表示サイズ, ...) というタプルのリスト
        columns = [column[0] for column in cursor.description]

        # 4. 2次元リストへ変換
        # fetchall() はタプルのリストを返すため、リスト内包表記で各行をリスト化します
        try:
            data_list = [list(row) for row in cursor.fetchall()]
        except Exception as e:
            raise Exception(f'データベースfetch中に予期せぬエラーです FetchTnju') from e
        finally:
            cursor.close()
            # cnxnは呼び出しもとでクローズ

        return columns, data_list


class FetchGrossWeight(IFetchDataForList):

    def __init__(self, cnxn) -> None:
        self.cnxn = cnxn
        

    def fetch_data(self) -> Tuple[List[str],List[List[Any]]]:

        cursor = self.cnxn.cursor()

        sqlQuery = ("SELECT MHINCD.HinHinCD AS 'hinban',"
                    " MHINCD.HinFree19 AS 'grossWeight'"
                    " From dbo.MHINCD"
                    " WHERE MHINCD.HinFree19 <> ''"
                    " ORDER BY MHINCD.HinHinCD"
                    )

        data_list: List[List[Any]] = []
        cursor.execute(sqlQuery)

        # 1. カラム名を取得（リスト内包表記）
        # cursor.description は (名前, 型, 表示サイズ, ...) というタプルのリスト
        columns = [column[0] for column in cursor.description]

        # 4. 2次元リストへ変換
        # fetchall() はタプルのリストを返すため、リスト内包表記で各行をリスト化します
        try:
            data_list = [list(row) for row in cursor.fetchall()]
        except Exception as e:
            raise Exception(f'データベースfetch中に予期せぬエラーです FetchGrossWeight') from e
        finally:
            cursor.close()
            # cnxnは呼び出しもとでクローズ

        return columns, data_list


class FetchUnsoutaiouHonsya(IFetchDataForList):

    def __init__(self, cnxn) -> None:
        self.cnxn = cnxn
        

    def fetch_data(self) -> Tuple[List[str],List[List[Any]]]:

        cursor = self.cnxn.cursor()

        sqlQuery = ("SELECT DesKojCD AS 'factoryCD',"
                    " DesTokCD AS 'tokuiCD',"
                    " DesNonyuCD AS 'nonyuCD',"
                    " DesLeadTime AS 'leadTime'," # int
                    " DesIsExport AS 'isExport'"  # int 1 or 0 
                    " From dbo.MDESTN_U2002"
                    " WHERE DesKojCD = '@0001'"
                    " ORDER BY DesKojCD,DesTokCD,DesNonyuCD"
                    )

        data_list: List[List[Any]] = []
        cursor.execute(sqlQuery)

        # 1. カラム名を取得（リスト内包表記）
        # cursor.description は (名前, 型, 表示サイズ, ...) というタプルのリスト
        columns = [column[0] for column in cursor.description]

        # 4. 2次元リストへ変換
        # fetchall() はタプルのリストを返すため、リスト内包表記で各行をリスト化します
        try:
            data_list = [list(row) for row in cursor.fetchall()]
        except Exception as e:
            raise Exception(f'データベースfetch中に予期せぬエラーです FetchUnsoutaiouHonsya') from e
        finally:
            cursor.close()
            # cnxnは呼び出しもとでクローズ

        return columns, data_list


class FetchUnsoutaiouToke(IFetchDataForList):

    def __init__(self, cnxn) -> None:
        self.cnxn = cnxn
        

    def fetch_data(self) -> Tuple[List[str],List[List[Any]]]:

        cursor = self.cnxn.cursor()

        sqlQuery = ("SELECT DesKojCD AS 'factoryCD',"
                    " DesTokCD AS 'tokuiCD',"
                    " DesNonyuCD AS 'nonyuCD',"
                    " DesLeadTime AS 'leadTime'," # int
                    " DesIsExport AS 'isExport'"  # int 1 or 0 
                    " From dbo.MDESTN_U2002"
                    " WHERE DesKojCD = '@0002'"
                    " ORDER BY DesKojCD,DesTokCD,DesNonyuCD"
                    )

        data_list: List[List[Any]] = []
        cursor.execute(sqlQuery)

        # 1. カラム名を取得（リスト内包表記）
        # cursor.description は (名前, 型, 表示サイズ, ...) というタプルのリスト
        columns = [column[0] for column in cursor.description]

        # 4. 2次元リストへ変換
        # fetchall() はタプルのリストを返すため、リスト内包表記で各行をリスト化します
        try:
            data_list = [list(row) for row in cursor.fetchall()]
        except Exception as e:
            raise Exception(f'データベースfetch中に予期せぬエラーです FetchUnsoutaiouToke') from e
        finally:
            cursor.close()
            # cnxnは呼び出しもとでクローズ

        return columns, data_list
        
    '''
        path = r'//192.168.1.247/共有/経理課ﾌｫﾙﾀﾞ/運賃計算関係/unsoutaiou_toke.csv'
        if platform.system() == 'Linux':
            path = r'/mnt/public/経理課ﾌｫﾙﾀﾞ/運賃計算関係/unsoutaiou_toke.csv'
        data = []
        try:
            with open(path, 'r', encoding='cp932') as f: 
                reader = csv.reader(f)
                data = [row for row in reader]
        except Exception as e:
            raise Exception('unsoutaiou_tokeのfetchに失敗です') from e

        columns = data[0]
        data_list = data[1:]
            
        return columns, data_list
        '''


class FetchSyukkaListCoa(IFetchDataForList):

    def __init__(self) -> None:
        pass
        

    def fetch_data(self) -> Tuple[List[Any],List[List[Any]]]:
        
        path = r'\\192.168.1.247\共有\営業課ﾌｫﾙﾀﾞ\櫻田\☆☆☆\売上処理(水野課長用)\出荷時添付リスト(20200731時点最新版).xlsx'
        if platform.system() == 'Linux':
            path = r'/mnt/public/営業課ﾌｫﾙﾀﾞ/櫻田/☆☆☆/売上処理(水野課長用)/出荷時添付リスト(20200731時点最新版).xlsx'
        data = []
        try:
            # data_only=True を足すだけで、数式ではなく計算結果の値を読み込めます
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb['成績表']

            # min_row=2 ２行目以降を取り込む iter_rowsタプルをリストに内包表記
            data = [list(row) for row in ws.iter_rows(min_row = 2, values_only=True)]

        except Exception as e:
            raise Exception('出荷時添付リスト(成績書)のfetchに失敗です') from e

        columns = data[0]
        data_list = data[1:]
            
        return columns, data_list


class FetchSyukkaListSiteiDenpyo(IFetchDataForList):

    def __init__(self) -> None:
        pass
        

    def fetch_data(self) -> Tuple[List[Any],List[List[Any]]]:
        
        path = r'\\192.168.1.247\共有\営業課ﾌｫﾙﾀﾞ\櫻田\☆☆☆\売上処理(水野課長用)\出荷時添付リスト(20200731時点最新版).xlsx'
        if platform.system() == 'Linux':
            path = r'/mnt/public/営業課ﾌｫﾙﾀﾞ/櫻田/☆☆☆/売上処理(水野課長用)/出荷時添付リスト(20200731時点最新版).xlsx'
        data = []
        try:
            # data_only=True を足すだけで、数式ではなく計算結果の値を読み込めます
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb['指定伝票']

            # min_row=2 ２行目以降を取り込む iter_rowsタプルをリストに内包表記
            data = [list(row) for row in ws.iter_rows(min_row = 2, values_only=True)]

        except Exception as e:
            raise Exception('出荷時添付リスト(指定伝票)のfetchに失敗です') from e

        columns = data[0]
        data_list = data[1:]
            
        return columns, data_list
